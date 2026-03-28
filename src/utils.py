"""
FIFA 2026 World Cup - Economic Impact Prediction
Utility Functions & Helpers
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.preprocessing import StandardScaler, MinMaxScaler
from scipy import stats

# ============================================================================
# DATA VALIDATION UTILITIES
# ============================================================================

def validate_dataframe(df, column_specs):
    """
    Validate dataframe against specifications
    
    Parameters:
    - df: DataFrame to validate
    - column_specs: dict with {column_name: {'type': type, 'min': min_val, 'max': max_val}}
    """
    issues = []
    
    for col, specs in column_specs.items():
        if col not in df.columns:
            issues.append(f"Missing column: {col}")
            continue
        
        # Check type
        if 'type' in specs:
            try:
                df[col] = df[col].astype(specs['type'])
            except:
                issues.append(f"Cannot convert {col} to {specs['type']}")
        
        # Check range
        if 'min' in specs:
            if df[col].min() < specs['min']:
                issues.append(f"{col} has values below minimum {specs['min']}")
        
        if 'max' in specs:
            if df[col].max() > specs['max']:
                issues.append(f"{col} has values above maximum {specs['max']}")
    
    return issues if issues else ["✓ All validations passed"]

def detect_outliers(df, columns, method='iqr', threshold=1.5):
    """
    Detect outliers using IQR or Z-score method
    """
    outliers = pd.DataFrame()
    
    for col in columns:
        if method == 'iqr':
            Q1 = df[col].quantile(0.25)
            Q3 = df[col].quantile(0.75)
            IQR = Q3 - Q1
            mask = (df[col] < Q1 - threshold*IQR) | (df[col] > Q3 + threshold*IQR)
        else:  # z-score
            z_scores = np.abs(stats.zscore(df[col].dropna()))
            mask = z_scores > threshold
        
        outliers[col] = mask
    
    return df[outliers.any(axis=1)]

# ============================================================================
# STATISTICAL ANALYSIS UTILITIES
# ============================================================================

def calculate_correlation_matrix(df, numeric_cols=None):
    """
    Calculate correlation matrix for numeric columns
    """
    if numeric_cols is None:
        numeric_cols = df.select_dtypes(include=[np.number]).columns
    
    return df[numeric_cols].corr()

def calculate_elasticity(df, dependent_var, independent_var):
    """
    Calculate elasticity: % change in Y / % change in X
    """
    pct_change_x = df[independent_var].pct_change().dropna()
    pct_change_y = df[dependent_var].pct_change().dropna()
    
    elasticity = pct_change_y / pct_change_x
    return elasticity.mean(), elasticity.std()

def comparison_vs_historical(metric_2026, historical_series):
    """
    Compare 2026 projection against historical average
    """
    historical_mean = historical_series.mean()
    historical_std = historical_series.std()
    
    z_score = (metric_2026 - historical_mean) / historical_std
    pct_diff = ((metric_2026 - historical_mean) / historical_mean) * 100
    
    return {
        'historical_mean': historical_mean,
        'historical_std': historical_std,
        'z_score': z_score,
        'pct_difference': pct_diff
    }

# ============================================================================
# SCALING & NORMALIZATION
# ============================================================================

def normalize_features(df, columns, method='minmax'):
    """
    Normalize specified columns
    
    Parameters:
    - method: 'minmax' or 'zscore'
    """
    df_normalized = df.copy()
    
    if method == 'minmax':
        scaler = MinMaxScaler()
    else:
        scaler = StandardScaler()
    
    df_normalized[columns] = scaler.fit_transform(df[columns])
    
    return df_normalized, scaler

def denormalize_features(scaled_values, scaler, columns):
    """
    Reverse normalization
    """
    return scaler.inverse_transform(scaled_values.reshape(-1, len(columns)))

# ============================================================================
# PROJECTION & FORECASTING UTILITIES
# ============================================================================

def project_with_growth_rate(base_value, growth_rate, periods):
    """
    Project values with compound growth rate
    """
    return base_value * (1 + growth_rate) ** periods

def project_with_inflation_adjustment(nominal_value, inflation_rates):
    """
    Adjust nominal value for inflation across multiple years
    """
    real_value = nominal_value
    for rate in inflation_rates:
        real_value = real_value / (1 + rate)
    return real_value

def create_confidence_intervals(point_estimate, std_error, confidence_level=0.95):
    """
    Create confidence intervals for estimates
    """
    z_score = stats.norm.ppf((1 + confidence_level) / 2)
    margin_of_error = z_score * std_error
    
    return {
        'lower_bound': point_estimate - margin_of_error,
        'point_estimate': point_estimate,
        'upper_bound': point_estimate + margin_of_error,
        'margin_of_error': margin_of_error
    }

# ============================================================================
# VISUALIZATION UTILITIES
# ============================================================================

def create_comparison_chart(historical_data, estimated_2026, metric_name):
    """
    Create comparison chart: Historical vs 2026 Estimate
    """
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))
    
    # Historical trend
    ax1.plot(historical_data.index, historical_data.values, marker='o', linewidth=2, markersize=8)
    ax1.axhline(y=historical_data.mean(), color='r', linestyle='--', label='Historical Avg')
    ax1.set_title(f'Historical {metric_name} Trend')
    ax1.set_xlabel('Year')
    ax1.set_ylabel(metric_name)
    ax1.legend()
    ax1.grid(True, alpha=0.3)
    
    # 2026 Comparison
    data = [historical_data.mean(), estimated_2026]
    colors = ['#1f77b4', '#ff7f0e']
    ax2.bar(['Historical Average', 'USA 2026 Estimate'], data, color=colors, alpha=0.7)
    ax2.set_title(f'{metric_name} Comparison')
    ax2.set_ylabel(metric_name)
    
    for i, v in enumerate(data):
        ax2.text(i, v + max(data)*0.02, f'{v:.0f}', ha='center', fontweight='bold')
    
    plt.tight_layout()
    return fig

def create_correlation_heatmap(correlation_matrix, title='Correlation Matrix'):
    """
    Create correlation heatmap
    """
    fig, ax = plt.subplots(figsize=(10, 8))
    
    sns.heatmap(correlation_matrix, annot=True, fmt='.2f', 
                cmap='coolwarm', center=0, square=True, 
                linewidths=1, cbar_kws={'label': 'Correlation'}, ax=ax)
    
    ax.set_title(title, fontsize=14, fontweight='bold')
    plt.tight_layout()
    
    return fig

def create_regional_impact_map(cities_df, impact_column):
    """
    Create regional impact visualization
    """
    fig, ax = plt.subplots(figsize=(12, 6))
    
    # Sort by impact
    sorted_data = cities_df.sort_values(by=impact_column, ascending=True)
    
    bars = ax.barh(range(len(sorted_data)), sorted_data[impact_column], color='steelblue')
    ax.set_yticks(range(len(sorted_data)))
    ax.set_yticklabels(sorted_data['City'])
    ax.set_xlabel(impact_column)
    ax.set_title(f'Estimated Economic Impact by City', fontsize=14, fontweight='bold')
    
    # Add value labels
    for i, (idx, row) in enumerate(sorted_data.iterrows()):
        ax.text(row[impact_column], i, f" ${row[impact_column]:.0f}M", va='center')
    
    ax.grid(axis='x', alpha=0.3)
    plt.tight_layout()
    
    return fig

# ============================================================================
# SCENARIO ANALYSIS
# ============================================================================

def create_scenarios(base_estimate, variations):
    """
    Create optimistic/pessimistic scenarios
    
    Parameters:
    - base_estimate: baseline projection
    - variations: dict with scenario names and % adjustments
    
    Example:
    variations = {
        'optimistic': 0.15,      # +15%
        'pessimistic': -0.10,    # -10%
        'conservative': -0.05    # -5%
    }
    """
    scenarios = {'base': base_estimate}
    
    for scenario_name, variation in variations.items():
        scenarios[scenario_name] = base_estimate * (1 + variation)
    
    return scenarios

def create_sensitivity_analysis(base_value, parameters_range):
    """
    Create sensitivity analysis table
    
    Parameters:
    - parameters_range: dict with {param_name: [min_val, max_val]}
    """
    sensitivity = {}
    
    for param, (min_val, max_val) in parameters_range.items():
        # Assume linear relationship for simplicity
        impact_low = base_value * (min_val / 100)
        impact_high = base_value * (max_val / 100)
        
        sensitivity[param] = {
            'low_scenario': impact_low,
            'base_estimate': base_value,
            'high_scenario': impact_high,
            'range': impact_high - impact_low
        }
    
    return sensitivity

# ============================================================================
# SUMMARY STATISTICS
# ============================================================================

def print_summary_stats(df, columns=None):
    """
    Print formatted summary statistics
    """
    if columns is None:
        columns = df.select_dtypes(include=[np.number]).columns
    
    stats_df = df[columns].describe().round(2)
    
    print("\n" + "="*70)
    print("SUMMARY STATISTICS")
    print("="*70)
    print(stats_df.to_string())
    print("="*70)

def export_summary_report(dataframes_dict, output_path):
    """
    Export multiple dataframes to Excel with summaries
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Format header
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
